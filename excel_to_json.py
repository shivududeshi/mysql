
from openpyxl import load_workbook

class ExcelToJson():
    def __init__(self,file_name):
        self.wb = load_workbook(file_name)
        self.data = {}
        self.conversion()

    def getJson(self):
        return self.data

    def conversion(self):
        ret = []
        try:
            ws = self.wb["Sheet1"]
            headers = ("ProductName","Description", "Addl_desc_product","prod_req_doc","prod_deliverables","prod_eta","Category","Price","Tax","Inclusicve ","MOU","Image_url","Var1_Name","var1_Desc","var1_required_docs","var1_deliverables","var1_ETA","var1_Price","Var2_Name","var2_Desc","var2_required_docs","var2_deliverables","var2_ETA","var2_Price","var3_name","var3_Desc","var3_required_docs","var3_deliverables","var3_ETA","var3_Price")
            for i in range(1,len(headers)+1):

                if ws.cell(1,i).value != headers[i-1]:
                    raise Exception(f'Unknown Data Format, expected {headers[i-1]}, found {ws.cell(1,i).value}')

            for row in ws.iter_rows(min_row=2):
                fn_name = row[0].value
                if type(fn_name) is  str and len(fn_name) > 0:
                    for item in ret:
                        if fn_name == item['display_name']:
                            raise Exception(f"Duplicate product {fn_name} found")
                    keys=("display_name","description","addl_description","req_doc","deliverables","eta","category","sale_val","tax_methods","tax_inclusicve","uom","image_url","variant_display_name","variant_description","variant_required_docs","variant_deliverables","variant_eta","variant_sale_val","variant_display_name","variant_description","variant_required_docs","variant_deliverables","variant_eta","variant_sale_val","variant_display_name","variant_description","variant_required_docs","variant_deliverables","variant_eta","variant_sale_val")
                    prod={}
                    prod['variants']=[{},{},{}]
                    for i in range(len(headers)):
                        if i<12:
                            prod[keys[i]]=row[i].value
                        elif i>=24:
                            prod['variants'][2][keys[i]]=row[i].value
                        elif i>=18:
                            prod['variants'][1][keys[i]]=row[i].value
                        elif i>=12:
                            prod['variants'][0][keys[i]]=row[i].value
                    ret.append(prod)
        except KeyError:
            pass
        except Exception as e:
            print('Error is :', e)
        # self.data['products'] = ret
        self.data = ret

if __name__=='__main__':

    o = ExcelToJson("MBG_products.xlsx")
    json = o.getJson()






